import os
import base64
import subprocess
from pathlib import Path
from datetime import date, datetime

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from supabase import Client, create_client

st.set_page_config(
    page_title="GoPropostas",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)

EDGE_FUNCTION_CREATE_SUBSCRIPTION_URL = "https://kwsnjozsfvhrddxycoco.supabase.co/functions/v1/create-subscription"
EDGE_FUNCTION_CREATE_PIX_URL = "https://kwsnjozsfvhrddxycoco.supabase.co/functions/v1/create-pix"
LOGO_PATH = "Apresentação de logo moderno e profissional.png"

# ---------------- VISUAL ----------------
def img_to_base64(path: str) -> str:
    if not Path(path).exists():
        return ""
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()

logo_base64 = img_to_base64(LOGO_PATH)

st.markdown("""
<style>
    .stApp {
        background: linear-gradient(180deg, #062B36 0%, #073846 55%, #0A4C5B 100%);
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #06232C 0%, #083845 100%);
        border-right: 1px solid rgba(255,255,255,0.08);
    }

    [data-testid="stSidebar"] * {
        color: #F4F7FA !important;
    }

    .block-container {
        padding-top: 1rem;
        padding-bottom: 2rem;
        max-width: 1380px;
    }

    .gp-topbar {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 18px;
        margin-bottom: 20px;
        padding: 18px 22px;
        border-radius: 24px;
        background: rgba(255,255,255,0.06);
        border: 1px solid rgba(255,255,255,0.08);
        backdrop-filter: blur(12px);
        box-shadow: 0 12px 32px rgba(0,0,0,0.18);
    }

    .gp-topbar img {
        height: 74px;
        width: auto;
        border-radius: 14px;
    }

    .gp-title {
        color: #F4F7FA;
        font-size: 2rem;
        font-weight: 800;
        margin: 0;
        line-height: 1.1;
    }

    .gp-subtitle {
        color: rgba(244,247,250,0.82);
        font-size: 0.98rem;
        margin-top: 4px;
    }

    .gp-card {
        background: #F8FBFD;
        border-radius: 24px;
        padding: 22px 24px;
        box-shadow: 0 12px 28px rgba(0,0,0,0.16);
        border: 1px solid rgba(12,109,132,0.10);
        margin-bottom: 18px;
    }

    .gp-card-dark {
        background: linear-gradient(135deg, #0A3D4B 0%, #0C6D84 100%);
        color: white;
        border-radius: 24px;
        padding: 22px 24px;
        box-shadow: 0 12px 28px rgba(0,0,0,0.22);
        margin-bottom: 18px;
    }

    .gp-section-title {
        color: #062B36;
        font-size: 1.15rem;
        font-weight: 800;
        margin-bottom: 14px;
    }

    .gp-card-dark .gp-section-title {
        color: white;
    }

    .gp-highlight {
        color: #F97316;
        font-weight: 800;
    }

    div[data-testid="stMetric"] {
        background: white;
        border-radius: 18px;
        padding: 14px 16px;
        border: 1px solid rgba(12,109,132,0.10);
        box-shadow: 0 8px 18px rgba(0,0,0,0.08);
    }

    div[data-testid="stMetricLabel"] {
        color: #0C6D84 !important;
        font-weight: 700 !important;
    }

    div[data-testid="stMetricValue"] {
        color: #062B36 !important;
        font-weight: 800 !important;
    }

    .stButton > button,
    .stDownloadButton > button {
        background: linear-gradient(90deg, #F97316 0%, #FF8E2B 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 14px !important;
        font-weight: 800 !important;
        min-height: 46px !important;
        box-shadow: 0 10px 18px rgba(249,115,22,0.28) !important;
    }

    .stButton > button:hover,
    .stDownloadButton > button:hover {
        transform: translateY(-1px);
        filter: brightness(1.03);
    }

    .stTextInput > div > div > input,
    .stNumberInput input,
    .stDateInput input,
    .stSelectbox div[data-baseweb="select"] > div,
    .stTextArea textarea,
    div[data-baseweb="input"] input,
    div[data-baseweb="base-input"] input {
        border-radius: 14px !important;
        border: 1px solid rgba(12,109,132,0.18) !important;
        background: #FFFFFF !important;
        color: #062B36 !important;
        -webkit-text-fill-color: #062B36 !important;
        caret-color: #062B36 !important;
    }

    .stTextInput > div > div > input::placeholder,
    .stNumberInput input::placeholder,
    .stDateInput input::placeholder,
    .stTextArea textarea::placeholder,
    div[data-baseweb="input"] input::placeholder,
    div[data-baseweb="base-input"] input::placeholder {
        color: #6B7C85 !important;
        opacity: 1 !important;
    }

    .stSelectbox div[data-baseweb="select"] * {
        color: #062B36 !important;
    }

    .stDateInput svg,
    .stSelectbox svg {
        fill: #062B36 !important;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }

    .stTabs [data-baseweb="tab"] {
        background: rgba(255,255,255,0.08);
        color: #F4F7FA;
        border-radius: 12px;
        padding: 10px 16px;
    }

    .stTabs [aria-selected="true"] {
        background: #F97316 !important;
        color: white !important;
    }

    hr {
        border-color: rgba(255,255,255,0.12);
    }

    .gp-mini {
        color: rgba(255,255,255,0.85);
        font-size: 0.95rem;
    }
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<div class="gp-topbar">
    {"<img src='data:image/png;base64," + logo_base64 + "'>" if logo_base64 else ""}
    <div>
        <div class="gp-title">GoPropostas</div>
        <div class="gp-subtitle">Sistema corporativo de propostas imobiliárias</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ---------------- SUPABASE LOGIN ----------------
@st.cache_resource
def get_supabase() -> Client:
    return create_client(
        st.secrets["SUPABASE_URL"].strip(),
        st.secrets["SUPABASE_KEY"].strip(),
    )

def buscar_profile_por_id(user_id: str):
    supabase = get_supabase()
    resp = (
        supabase.table("profiles")
        .select("*")
        .eq("id", user_id)
        .limit(1)
        .execute()
    )
    return resp.data[0] if resp.data else None

def buscar_profile_por_email(email: str):
    supabase = get_supabase()
    resp = (
        supabase.table("profiles")
        .select("*")
        .eq("email", email)
        .limit(1)
        .execute()
    )
    return resp.data[0] if resp.data else None

def buscar_assinatura(user_id: str):
    supabase = get_supabase()
    resp = (
        supabase.table("assinaturas")
        .select("*")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
        .limit(1)
        .execute()
    )
    return resp.data[0] if resp.data else None

def assinatura_ativa_para_acesso(assinatura: dict) -> bool:
    if not assinatura:
        return False
    if not assinatura.get("assinatura_ativa"):
        return False

    proximo = assinatura.get("proximo_cobranca_em")
    if not proximo:
        return bool(assinatura.get("assinatura_ativa"))

    try:
        proximo_dt = datetime.fromisoformat(str(proximo).replace("Z", "+00:00"))
        agora = datetime.now(proximo_dt.tzinfo) if proximo_dt.tzinfo else datetime.now()
        return proximo_dt >= agora
    except Exception:
        return bool(assinatura.get("assinatura_ativa"))

def criar_assinatura_mp(user_id: str, email: str):
    headers = {
        "Content-Type": "application/json",
    }
    payload = {
        "user_id": user_id,
        "email": email,
    }

    resp = requests.post(
        EDGE_FUNCTION_CREATE_SUBSCRIPTION_URL,
        json=payload,
        headers=headers,
        timeout=30,
    )

    try:
        return resp.json()
    except Exception:
        return {
            "error": f"Resposta inválida da função: status {resp.status_code}",
            "raw_text": resp.text,
        }

def criar_pix(user_id: str, email: str):
    headers = {
        "Content-Type": "application/json",
    }
    payload = {
        "user_id": user_id,
        "email": email,
    }

    resp = requests.post(
        EDGE_FUNCTION_CREATE_PIX_URL,
        json=payload,
        headers=headers,
        timeout=30,
    )

    try:
        return resp.json()
    except Exception:
        return {
            "error": f"Resposta inválida da função PIX: status {resp.status_code}",
            "raw_text": resp.text,
        }

def login_com_supabase(email: str, senha: str):
    supabase = get_supabase()
    return supabase.auth.sign_in_with_password({
        "email": email,
        "password": senha,
    })

def cadastrar_com_supabase(nome: str, email: str, senha: str):
    supabase = get_supabase()
    return supabase.auth.sign_up({
        "email": email,
        "password": senha,
        "options": {
            "data": {
                "nome": nome
            }
        }
    })

def init_auth_state():
    defaults = {
        "logado": False,
        "usuario_id": "",
        "usuario_email": "",
        "usuario_nome": "",
        "tipo": "",
        "sb_access_token": "",
        "sb_refresh_token": "",
    }
    for chave, valor in defaults.items():
        if chave not in st.session_state:
            st.session_state[chave] = valor

def aplicar_login(profile: dict):
    st.session_state["logado"] = True
    st.session_state["usuario_id"] = profile["id"]
    st.session_state["usuario_email"] = profile["email"]
    st.session_state["usuario_nome"] = profile.get("nome") or profile["email"]
    st.session_state["tipo"] = profile["tipo"]

def salvar_tokens_da_sessao(auth_response):
    session = getattr(auth_response, "session", None)
    if session:
        st.session_state["sb_access_token"] = session.access_token or ""
        st.session_state["sb_refresh_token"] = session.refresh_token or ""

def tentar_restaurar_sessao():
    if st.session_state.get("logado"):
        return

    access_token = st.session_state.get("sb_access_token", "")
    refresh_token = st.session_state.get("sb_refresh_token", "")

    if not access_token or not refresh_token:
        return

    try:
        supabase = get_supabase()
        supabase.auth.set_session(access_token, refresh_token)
        sessao = supabase.auth.get_session()
        session = getattr(sessao, "session", None)

        if not session or not session.user:
            return

        profile = buscar_profile_por_id(session.user.id)
        if profile:
            aplicar_login(profile)
    except Exception:
        pass

def tela_login():
    st.markdown('<div class="gp-card-dark">', unsafe_allow_html=True)
    st.title("🔐 Entrar no sistema")
    st.markdown('<div class="gp-mini">Acesse ou crie sua conta para continuar.</div>', unsafe_allow_html=True)

    abas = st.tabs(["Login", "Criar conta"])

    with abas[0]:
        email = st.text_input("Email", key="login_email")
        senha = st.text_input("Senha", type="password", key="login_senha")

        if st.button("Entrar", key="btn_login", use_container_width=True):
            try:
                resp = login_com_supabase(email, senha)
                user = resp.user

                if not user:
                    st.error("Email ou senha inválidos.")
                    return

                salvar_tokens_da_sessao(resp)

                profile = buscar_profile_por_id(user.id)
                if not profile:
                    st.error("Perfil não encontrado. Verifique se o trigger do Supabase foi criado.")
                    return

                aplicar_login(profile)
                st.success("Login realizado com sucesso!")
                st.rerun()

            except Exception as e:
                st.error(f"Erro no login: {e}")

    with abas[1]:
        nome_cadastro = st.text_input("Nome completo", key="cad_nome")
        email_cadastro = st.text_input("Email", key="cad_email")
        senha_cadastro = st.text_input("Senha", type="password", key="cad_senha")
        confirmar = st.text_input("Confirmar senha", type="password", key="cad_confirm")

        if st.button("Criar conta", key="btn_cadastro", use_container_width=True):
            if senha_cadastro != confirmar:
                st.warning("Senhas não conferem.")
                return
            if not nome_cadastro.strip() or not email_cadastro.strip() or not senha_cadastro.strip():
                st.warning("Preencha todos os campos.")
                return

            try:
                existente = buscar_profile_por_email(email_cadastro)
                if existente:
                    st.warning("Já existe uma conta com esse email.")
                    return

                resp = cadastrar_com_supabase(nome_cadastro, email_cadastro, senha_cadastro)
                user = resp.user

                if user:
                    st.success("Conta criada com sucesso! Agora faça login.")
                else:
                    st.success("Conta criada! Verifique seu email para confirmar o cadastro antes de entrar.")

            except Exception as e:
                st.error(f"Erro ao criar conta: {e}")

    st.markdown("</div>", unsafe_allow_html=True)

def logout():
    if st.sidebar.button("🚪 Sair", key="logout", use_container_width=True):
        try:
            get_supabase().auth.sign_out()
        except Exception:
            pass

        for chave in [
            "logado", "usuario_id", "usuario_email", "usuario_nome", "tipo",
            "sb_access_token", "sb_refresh_token"
        ]:
            if chave in st.session_state:
                del st.session_state[chave]

        st.rerun()

# ---------------- CONTROLE LOGIN ----------------
init_auth_state()
tentar_restaurar_sessao()

if not st.session_state["logado"]:
    tela_login()
    st.stop()

assinatura = buscar_assinatura(st.session_state["usuario_id"])
eh_admin = st.session_state.get("tipo") == "admin"
acesso_liberado = eh_admin or assinatura_ativa_para_acesso(assinatura)

if not acesso_liberado:
    st.markdown("""
    <div class="gp-card-dark">
        <div style="font-size:1.5rem;font-weight:800;">💳 Assinatura GoPropostas</div>
        <div style="margin-top:8px;font-size:1rem;opacity:0.92;">
            Escolha a melhor forma para acessar o sistema:
        </div>
        <div style="margin-top:16px;font-size:1.95rem;font-weight:900;">
            R$ 15,00 <span style="font-size:1rem;font-weight:500;">/ mês</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if assinatura:
        st.info(f"Status atual: {assinatura.get('status', 'pendente')}")
        if assinatura.get("proximo_cobranca_em"):
            st.caption(f"Validade / próxima cobrança: {assinatura.get('proximo_cobranca_em')}")

    col_pag_1, col_pag_2 = st.columns(2)

    with col_pag_1:
        st.markdown('<div class="gp-card">', unsafe_allow_html=True)
        st.markdown('<div class="gp-section-title">💳 Assinatura no cartão</div>', unsafe_allow_html=True)
        st.write("Cobrança recorrente automática mensal.")
        if st.button("Assinar no cartão", use_container_width=True):
            try:
                data = criar_assinatura_mp(
                    st.session_state["usuario_id"],
                    st.session_state["usuario_email"]
                )
                link = data.get("init_point") or data.get("sandbox_init_point")

                if link:
                    st.link_button("👉 Ir para pagamento", link, use_container_width=True)
                else:
                    st.error(f"Erro ao gerar link de pagamento: {data}")
            except Exception as e:
                st.error(f"Erro ao iniciar assinatura: {e}")
        st.markdown('</div>', unsafe_allow_html=True)

    with col_pag_2:
        st.markdown('<div class="gp-card">', unsafe_allow_html=True)
        st.markdown('<div class="gp-section-title">🧾 PIX mensal</div>', unsafe_allow_html=True)
        st.write("Pague manualmente via PIX e tenha acesso por 30 dias.")
        if st.button("Gerar PIX", use_container_width=True):
            try:
                data = criar_pix(
                    st.session_state["usuario_id"],
                    st.session_state["usuario_email"]
                )

                if "qr_code_base64" in data:
                    st.image(f"data:image/png;base64,{data['qr_code_base64']}")
                    st.code(data["qr_code"])
                    st.success("Escaneie ou copie o PIX.")
                else:
                    st.error(f"Erro ao gerar PIX: {data}")
            except Exception as e:
                st.error(f"Erro ao gerar PIX: {e}")
        st.markdown('</div>', unsafe_allow_html=True)

    st.stop()

st.sidebar.write(f"👤 {st.session_state['usuario_nome']}")
st.sidebar.write(f"📧 {st.session_state['usuario_email']}")
st.sidebar.write(f"🔑 {st.session_state['tipo']}")
logout()

# ---------------- EMPREENDIMENTOS ----------------
empreendimentos = {
    "Frei Galvão": {
        "proprietario": "Frei Galvão empreendimentos imobiliários",
        "nome": "Loteamento Frei Galvão",
        "logradouro": "Avenida Fazenda Bananal",
        "tabela": "tabela_frei_galvao.xlsx",
    }
}

# ---------------- UTILITÁRIOS ----------------
@st.cache_data
def carregar_tabela(arquivo):
    df = pd.read_excel(arquivo, skiprows=11)
    df.columns = df.columns.str.strip().str.lower()
    return df

def limpar(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except Exception:
        return 0.0

def buscar(linha, nomes):
    for col in linha.index:
        for nome in nomes:
            if nome.lower() in col.lower():
                return limpar(linha[col])
    return 0.0

def excel_para_pdf(arquivo):
    subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "pdf", arquivo],
        check=False,
    )
    return arquivo.replace(".xlsx", ".pdf")

def calcular_idade_em_data(nascimento: date, data_referencia: date) -> int:
    return data_referencia.year - nascimento.year - (
        (data_referencia.month, data_referencia.day) < (nascimento.month, nascimento.day)
    )

def adicionar_meses(data_base: date, meses: int) -> date:
    ano = data_base.year + (data_base.month - 1 + meses) // 12
    mes = (data_base.month - 1 + meses) % 12 + 1
    ultimo_dia = [
        31,
        29 if (ano % 4 == 0 and (ano % 100 != 0 or ano % 400 == 0)) else 28,
        31, 30, 31, 30, 31, 31, 30, 31, 30, 31
    ][mes - 1]
    dia = min(data_base.day, ultimo_dia)
    return date(ano, mes, dia)

# ---------------- EXCEL ----------------
def preencher_proposta(d, modelo="modelo_proposta.xlsx"):
    wb = load_workbook(modelo)
    ws = wb.active

    ws["E5"] = d["nome"]
    ws["D6"] = d["cpf"]
    ws["J6"] = d["telefone"]
    ws["O6"] = d["fixo"]
    ws["D7"] = d["nacionalidade"]
    ws["J7"] = d["profissao"]
    ws["P7"] = d["fone_pref"]
    ws["D8"] = d["estado_civil"]
    ws["O8"] = d["renda"]
    ws["E9"] = d["email"]

    ws["G11"] = d["conjuge"]
    ws["D13"] = d["cpf2"]
    ws["J13"] = d["tel2"]
    ws["O13"] = d["fixo2"]
    ws["D14"] = d["nac2"]
    ws["J14"] = d["prof2"]
    ws["P14"] = d["fone2"]
    ws["D15"] = d["civil2"]
    ws["O15"] = d["renda2"]

    ws["G18"] = d["proprietario"]
    ws["G19"] = d["empreendimento"]
    ws["C20"] = d["logradouro"]
    ws["I20"] = d["unidade"]
    ws["Q20"] = d["area"]

    ws["C21"] = d["valor_negocio"]
    ws["J21"] = d["entrada_total"]
    ws["O21"] = d["valor_imovel"]

    ws["B24"] = 1
    ws["C24"] = d["entrada_imovel"]
    ws["G24"] = "Única"
    ws["K24"] = d["data_venc_emp"]

    ws["B25"] = "36x"
    ws["C25"] = d["parcela_36"]
    ws["G25"] = "Mensal"
    ws["K25"] = d["data_parcelas"]

    ws["B26"] = 1
    ws["C26"] = d["saldo"]
    ws["G26"] = "Única"
    ws["K26"] = d["data_saldo"]

    ws["P24"] = "Fixo"
    ws["P25"] = "Reajustável"
    ws["P26"] = "Reajustável"

    ws["B33"] = 1
    ws["C33"] = d["entrada_cliente"]
    ws["G33"] = "Única"
    ws["P33"] = "À vista"
    ws["K33"] = d["data_ato"]
    ws["K33"].alignment = Alignment(horizontal="center", vertical="center")

    if d["entrada_quitada"]:
        ws["B34"] = ""
        ws["C34"] = ""
        ws["G34"] = ""
        ws["P34"] = ""
        ws["K34"] = ""

        ws["B35"] = ""
        ws["C35"] = ""
        ws["G35"] = ""
        ws["P35"] = ""
        ws["K35"] = ""
    else:
        ws["B34"] = d["parcelas_iguais"]
        ws["C34"] = d["valor_parcela_igual"]
        ws["G34"] = "Mensal" if d["parcelas_iguais"] > 0 else ""
        ws["P34"] = "Fixo"
        ws["K34"] = d["data_parc_entrada"]
        ws["K34"].alignment = Alignment(horizontal="center", vertical="center")

        if d["usar_diferente"]:
            ws["B35"] = 1
            ws["C35"] = d["parcela_diferente"]
            ws["G35"] = "Única"
            ws["P35"] = "Fixa"
            ws["K35"] = d["data_parcela_diferente_manual"]

            for cel in ["B35", "G35", "K35", "P35"]:
                ws[cel].alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws["B35"] = ""
            ws["C35"] = ""
            ws["G35"] = ""
            ws["P35"] = ""
            ws["K35"] = ""

    arquivo = "proposta.xlsx"
    wb.save(arquivo)
    return arquivo

# ---------------- APP ----------------
st.markdown('<div class="gp-card"><div class="gp-section-title">🏢 Empreendimento</div>', unsafe_allow_html=True)

emp_nome = st.selectbox("Selecione", list(empreendimentos.keys()), key="emp")
emp = empreendimentos[emp_nome]

df = carregar_tabela(emp["tabela"])
col = df.columns[0]
unidade = st.selectbox("Lote", df[col].dropna().unique(), key="lote")
linha = df[df[col] == unidade].iloc[0]

valor_negocio = buscar(linha, ["valor negócio"])
entrada_imovel = buscar(linha, ["entrada imovel"])
intermed = buscar(linha, ["intermediação"])
parcela_36 = buscar(linha, ["36x"])
saldo = buscar(linha, ["saldo"])
area = buscar(linha, ["área", "area"])
valor_imovel = buscar(linha, ["valor imóvel"])

entrada_total = intermed + entrada_imovel
ato_min = valor_negocio * 0.003

st.markdown('</div>', unsafe_allow_html=True)

col_form_1, col_form_2 = st.columns(2)

with col_form_1:
    st.markdown('<div class="gp-card"><div class="gp-section-title">👤 Cliente</div>', unsafe_allow_html=True)
    nome = st.text_input("Nome", key="nome")
    cpf = st.text_input("CPF", key="cpf")
    telefone = st.text_input("Telefone", key="tel")
    fixo = st.text_input("Fixo", key="fixo")
    nacionalidade = st.text_input("Nacionalidade", key="nac")
    profissao = st.text_input("Profissão", key="prof")
    fone_pref = st.text_input("Fone preferência", key="fonepref")
    estado_civil = st.text_input("Estado civil", key="civil")
    renda = st.text_input("Renda", key="renda")
    email = st.text_input("Email", key="email")
    data_nascimento = st.date_input(
        "Data de nascimento do cliente",
        value=date(1980, 1, 1),
        min_value=date(1900, 1, 1),
        max_value=date.today(),
        key="data_nascimento"
    )
    st.markdown('</div>', unsafe_allow_html=True)

with col_form_2:
    st.markdown('<div class="gp-card"><div class="gp-section-title">👫 Cônjuge</div>', unsafe_allow_html=True)
    conjuge = st.text_input("Nome", key="conj")
    cpf2 = st.text_input("CPF", key="cpf2")
    tel2 = st.text_input("Telefone", key="tel2")
    fixo2 = st.text_input("Fixo", key="fixo2")
    nac2 = st.text_input("Nacionalidade", key="nac2")
    prof2 = st.text_input("Profissão", key="prof2")
    fone2 = st.text_input("Fone preferência", key="fone2")
    civil2 = st.text_input("Estado civil", key="civil2")
    renda2 = st.text_input("Renda", key="renda2")
    st.markdown('</div>', unsafe_allow_html=True)

col_data_1, col_data_2 = st.columns(2)

with col_data_1:
    st.markdown('<div class="gp-card"><div class="gp-section-title">📅 Datas de Vencimento</div>', unsafe_allow_html=True)
    data_venc_emp = st.date_input("Data Vencimento Empreendedor", key="venc_emp")
    data_parcelas = st.date_input("Data Parcelas", key="venc_parc")
    data_saldo = st.date_input("Data Saldo Devedor", key="venc_saldo")
    st.markdown('</div>', unsafe_allow_html=True)

with col_data_2:
    st.markdown('<div class="gp-card"><div class="gp-section-title">📅 Datas da Entrada</div>', unsafe_allow_html=True)
    data_ato = st.date_input("Data do ato", key="data_ato")
    data_parc_entrada = st.date_input("Data primeiras parcelas entrada", key="data_parc_entrada")
    data_parc_diferente = st.date_input("Data da parcela diferente", key="data_parc_dif")
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="gp-card"><div class="gp-section-title">💰 Condições</div>', unsafe_allow_html=True)

valor_cliente = st.number_input("Entrada cliente", min_value=0.0, key="entrada")
personalizar = st.checkbox("⚙️ Personalizar", key="pers")

ato_manual = st.number_input("Valor ato", min_value=0.0, key="ato_manual") if personalizar else 0
ato = ato_manual if ato_manual > 0 else ato_min

restante = entrada_total - valor_cliente
if restante < 0:
    restante = 0

entrada_quitada = restante <= 0.01

valor_minimo_entrada = ato_min
erros_validacao = []
avisos_validacao = []

if valor_cliente <= 0:
    avisos_validacao.append("Nenhum valor foi informado em Entrada cliente.")

if valor_cliente < valor_minimo_entrada:
    erros_validacao.append(
        f"Entrada cliente menor que o mínimo. Mínimo recomendado: R$ {valor_minimo_entrada:,.2f}"
    )

if valor_cliente > entrada_total:
    avisos_validacao.append("Entrada cliente maior que a entrada total. O excedente não será parcelado.")

parcelas = st.slider("Parcelas", 1, 4, 1, key="parc")

usar_diferente = False
parcela_diferente = 0
data_parcela_diferente = ""
parcelas_iguais = 0
valor_parcela_igual = 0

if not entrada_quitada:
    parcelas_iguais = parcelas
    valor_parcela_igual = restante / parcelas if parcelas > 0 else 0

if personalizar and parcelas > 1 and not entrada_quitada:
    parcela_editada = st.number_input("Parcela diferente", min_value=0.0, key="diff")
    restante_auto = restante - parcela_editada

    if restante_auto < 0:
        restante_auto = 0
        avisos_validacao.append("A parcela diferente está maior que o restante disponível.")

    valor_parcela_igual = restante_auto / (parcelas - 1)

    if abs(parcela_editada - valor_parcela_igual) > 0.01:
        usar_diferente = True
        parcela_diferente = parcela_editada
        parcelas_iguais = parcelas - 1
        data_parcela_diferente = st.date_input("Data parcela diferente", key="data_diff")

        if parcela_diferente > restante:
            erros_validacao.append("A parcela diferente não pode ser maior que o restante da entrada.")

st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="gp-card"><div class="gp-section-title">🏡 Detalhes do Lote</div>', unsafe_allow_html=True)
col_l1, col_l2 = st.columns(2)
with col_l1:
    st.metric("Unidade", unidade)
    st.metric("Área (m²)", f"{area:.2f}")
    st.metric("Entrada Imóvel", f"R$ {entrada_imovel:,.2f}")

with col_l2:
    st.metric("Valor Negócio", f"R$ {valor_negocio:,.2f}")
    st.metric("Valor Imóvel", f"R$ {valor_imovel:,.2f}")
    st.metric("Intermediação", f"R$ {intermed:,.2f}")
st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="gp-card"><div class="gp-section-title">📊 Painel de Cálculo</div>', unsafe_allow_html=True)
col_c1, col_c2 = st.columns(2)
with col_c1:
    st.metric("Entrada Total", f"R$ {entrada_total:,.2f}")
    st.metric("Entrada Cliente (C33)", f"R$ {valor_cliente:,.2f}")
    st.metric("Valor mínimo", f"R$ {valor_minimo_entrada:,.2f}")

with col_c2:
    st.metric("Ato informado", f"R$ {ato:,.2f}")
    st.metric("Restante para parcelar", f"R$ {restante:,.2f}")
    st.metric("Quantidade de parcelas", f"{parcelas}")

st.markdown("### 📅 Parcelamento da entrada")
if entrada_quitada:
    st.success("Entrada paga à vista")
elif parcelas > 1:
    if usar_diferente:
        st.info(
            f"{parcelas_iguais}x de R$ {valor_parcela_igual:,.2f} + "
            f"1x de R$ {parcela_diferente:,.2f}"
        )
    else:
        st.success(f"{parcelas}x de R$ {valor_parcela_igual:,.2f}")
else:
    st.success("Pagamento em parcela única")

if avisos_validacao:
    for aviso in avisos_validacao:
        st.warning(f"⚠️ {aviso}")

if erros_validacao:
    for erro in erros_validacao:
        st.error(f"❌ {erro}")

st.markdown('</div>', unsafe_allow_html=True)

proposta_pode_ser_gerada = len(erros_validacao) == 0

if st.button("Gerar Proposta", use_container_width=True, disabled=not proposta_pode_ser_gerada):
    data_final_36_parcelas = adicionar_meses(data_parcelas, 36)
    idade_apos_36 = calcular_idade_em_data(data_nascimento, data_final_36_parcelas)

    if idade_apos_36 >= 75:
        st.warning("Cliente não conseguirá refinanciar após as 36 parcelas (idade superior a 75 anos)")

    dados = {
        "nome": nome,
        "cpf": cpf,
        "telefone": telefone,
        "fixo": fixo,
        "nacionalidade": nacionalidade,
        "profissao": profissao,
        "fone_pref": fone_pref,
        "estado_civil": estado_civil,
        "renda": renda,
        "email": email,
        "conjuge": conjuge,
        "cpf2": cpf2,
        "tel2": tel2,
        "fixo2": fixo2,
        "nac2": nac2,
        "prof2": prof2,
        "fone2": fone2,
        "civil2": civil2,
        "renda2": renda2,
        "proprietario": emp["proprietario"],
        "empreendimento": emp["nome"],
        "logradouro": emp["logradouro"],
        "unidade": unidade,
        "area": area,
        "valor_negocio": valor_negocio,
        "entrada_total": entrada_total,
        "valor_imovel": valor_imovel,
        "entrada_imovel": entrada_imovel,
        "parcela_36": parcela_36,
        "saldo": saldo,
        "entrada_cliente": valor_cliente,
        "entrada_quitada": entrada_quitada,
        "ato": ato,
        "parcelas_iguais": parcelas_iguais,
        "valor_parcela_igual": valor_parcela_igual,
        "usar_diferente": usar_diferente,
        "parcela_diferente": parcela_diferente,
        "data_parcela_diferente": data_parcela_diferente.strftime("%d/%m/%Y") if usar_diferente else "",
        "data_venc_emp": data_venc_emp.strftime("%d/%m/%Y"),
        "data_parcelas": data_parcelas.strftime("%d/%m/%Y"),
        "data_saldo": data_saldo.strftime("%d/%m/%Y"),
        "data_ato": data_ato.strftime("%d/%m/%Y") if data_ato else "",
        "data_parc_entrada": data_parc_entrada.strftime("%d/%m/%Y") if data_parc_entrada else "",
        "data_parcela_diferente_manual": data_parc_diferente.strftime("%d/%m/%Y") if data_parc_diferente else "",
    }

    excel = preencher_proposta(dados)
    pdf = excel_para_pdf(excel)

    st.success("✅ Proposta gerada com sucesso!")

    col_download_1, col_download_2 = st.columns(2)

    with col_download_1:
        with open(pdf, "rb") as f_pdf:
            st.download_button(
                "📥 Baixar PDF",
                f_pdf,
                file_name=f"Proposta_{unidade}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

    with col_download_2:
        with open(excel, "rb") as f_excel:
            st.download_button(
                "📥 Baixar Excel",
                f_excel,
                file_name=f"Proposta_{unidade}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )